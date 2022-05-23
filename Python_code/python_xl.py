# Libraries (requires openpyxl)
from openpyxl import Workbook
from openpyxl import load_workbook

# Loading table
sourcefile = "Avista_Meter_Top.xlsx"
wb=load_workbook(sourcefile)

# initialize variables
cap = [None]*62
res = [None]*62
other = [None]*62
cap_count = 1
res_count = 1
other_count = 1

# store columns in lists. These lists are compatible with K1830
ws = wb.active
feederid_column = ws['B']
skip_column = ws['C']
footprints_column = ws['G']
nozzle_column = ws['I']
comp_footprint_col = ws['D']
comp_nozzlenumber_col = ws['F']

# Create the list
feederid = [cell.value for cell in feederid_column[1:63]]
skip = [cell.value for cell in skip_column[1:63]]
footprints = [cell.value for cell in footprints_column[1:63]]
nozzle = [cell.value for cell in nozzle_column[1:63]]
comp_feederid = [cell.value for cell in feederid_column[85:]]
comp_footprint = [cell.value for cell in comp_footprint_col[85:]]
comp_nozzlenumber = [cell.value for cell in comp_nozzlenumber_col[85:]]

# Check if lists are correct
# print("feederid:",feederid,"\n")
# print("skip:",skip,"\n")
# print("footprints:",footprints,"\n")
# print("nozzle:",nozzle,"\n")
# print("comp_footprint:",comp_footprint,"\n")
# print("comp_nozzlenumber:",comp_nozzlenumber,"\n")

k = 0
for i in range(62):
    # ignore components with skip set to "YES"
    if skip[i] == "YES":
        k+=1
        nozzle[i] = None
        continue
    if footprints[i].endswith("_C"):
        cap[i] = footprints[i]
        nozzle[i] = cap_count
        cap_count += 1
        if cap_count > 8:
            cap_count = 1
    elif footprints[i].endswith("_R"):
        res[i] = footprints[i]
        nozzle[i] = res_count
        res_count += 1
        if res_count > 8:
            res_count = 1       
    else:
        other[i] = footprints[i]
        nozzle[i] = other_count
        other_count += 1
        if other_count > 8:
            other_count = 1

for r in range(2,64):
    ws.cell(row=r,column=9).value=nozzle[r-2]

for i in range(62):
    for j in range(0,len(comp_nozzlenumber)):
        if feederid[i] == comp_feederid[j]:
            comp_nozzlenumber[j] = nozzle[i]

# for j in range(0,len(comp_nozzlenumber)):
#     print(comp_nozzlenumber[j])

index = 0
for r in range(86,86+len(comp_nozzlenumber)):
    # print(r)
    ws.cell(row=r,column=6).value=comp_nozzlenumber[index]
    # print(comp_nozzlenumber[index])
    index += 1

wb.save('Avista_Meter_Top.xlsx')

